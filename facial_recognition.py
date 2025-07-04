import face_recognition
import cv2
import numpy as np
from picamera2 import Picamera2
import time
import pickle
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime 

# Load known face encodings
print("[INFO] Loading encodings...")
with open("encodings.pickle", "rb") as f:
    data = pickle.loads(f.read())
known_face_encodings = data["encodings"]
known_face_names = data["names"]

# Initialize the camera
picam2 = Picamera2()
picam2.configure(picam2.create_preview_configuration(main={"format": 'XRGB8888', "size": (1920, 1080)}))
picam2.start()

# Define daily attendance file path
def get_today_attendance_file():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"attendance_{date_str}.xlsx"

# Create new attendance file if it doesn't exist
def ensure_attendance_file():
    file_path = get_today_attendance_file()
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Date", "Time"])
        wb.save(file_path)

# Mark attendance and return whether it's already marked
def mark_attendance(name):
    file_path = get_today_attendance_file()
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    wb = load_workbook(file_path)
    ws = wb.active

    # Check if already marked
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name:
            return True  # Already marked

    ws.append([name, date_str, time_str])
    wb.save(file_path)
    print(f"[ATTENDANCE] {name} marked present at {time_str}")
    return False

# Global state
cv_scaler = 4
face_locations = []
face_encodings = []
face_names = []
face_confidences = []
already_marked_flags = []
frame_count = 0
start_time = time.time()
fps = 0

# Ensure today's Excel file is created
ensure_attendance_file()

def process_frame(frame):
    global face_locations, face_encodings, face_names, face_confidences, already_marked_flags

    resized_frame = cv2.resize(frame, (0, 0), fx=1/cv_scaler, fy=1/cv_scaler)
    rgb_resized_frame = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_resized_frame)
    face_encodings = face_recognition.face_encodings(rgb_resized_frame, face_locations, model='large')

    face_names = []
    face_confidences = []
    already_marked_flags = []

    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
        confidence = 0.0

        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]
            confidence = (1.0 - face_distances[best_match_index]) * 100

        if name != "Unknown":
            already_marked = mark_attendance(name)
        else:
            already_marked = False

        face_names.append(name)
        face_confidences.append(confidence)
        already_marked_flags.append(already_marked)

    return frame

def draw_results(frame):
    for (top, right, bottom, left), name, confidence, marked in zip(face_locations, face_names, face_confidences, already_marked_flags):
        top *= cv_scaler
        right *= cv_scaler
        bottom *= cv_scaler
        left *= cv_scaler

        label = f"{name} ({confidence:.1f}%)"
        
        # Draw face box
        cv2.rectangle(frame, (left, top), (right, bottom), (244, 42, 3), 3)

        # Draw name and confidence above
        cv2.rectangle(frame, (left - 3, top - 35), (right + 3, top), (244, 42, 3), cv2.FILLED)
        cv2.putText(frame, label, (left + 6, top - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)

        # Draw subtle "Marked" tag below
        if marked and name != "Unknown":
            cv2.putText(frame, "Marked", (left + 14, bottom + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (200, 255, 200), 1)

    return frame

def calculate_fps():
    global frame_count, start_time, fps
    frame_count += 1
    elapsed_time = time.time() - start_time
    if elapsed_time > 1:
        fps = frame_count / elapsed_time
        frame_count = 0
        start_time = time.time()
    return fps

# Main loop
while True:
    frame = picam2.capture_array()
    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)

    processed_frame = process_frame(frame.copy())
    display_frame = draw_results(frame)

    current_fps = calculate_fps()
    cv2.putText(display_frame, f"FPS: {current_fps:.1f}", (display_frame.shape[1] - 150, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

    cv2.imshow('Attendance System', display_frame)

    if cv2.waitKey(1) == ord("q"):
        break

cv2.destroyAllWindows()
picam2.stop()
